"""Recalculate workbook scenarios and compare them with the generated runtime."""

from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domain.models import (
    ParityRun,
    ParityScenario,
    ParityStatus,
    QuoteCreateRequest,
    SystemBlueprint,
)
from app.parity.scenarios import golden_scenarios
from app.runtime.service import QuoteRuntime, RuntimeValidationError


OUTPUT_COLUMNS = {
    "unit_price": "F",
    "revenue": "G",
    "cost": "H",
    "gross_margin": "I",
    "approval_status": "J",
}
INPUT_COLUMNS = {
    "client_id": "A",
    "product_id": "C",
    "quantity": "D",
    "discount": "E",
}
TOLERANCES = {
    "unit_price": 0.01,
    "revenue": 0.01,
    "cost": 0.01,
    "gross_margin": 0.000001,
}


def find_soffice() -> str | None:
    configured = os.getenv("SOFFICE_BIN", "").strip()
    return configured or shutil.which("soffice") or shutil.which("libreoffice")


def _normalise_value(value: Any) -> Any:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    return str(value)


def _recalculate_case(soffice: str, workbook_path: Path, inputs: dict[str, Any]) -> dict[str, Any]:
    with tempfile.TemporaryDirectory(prefix="sheet-parity-") as directory:
        workspace = Path(directory)
        input_path = workspace / "case.xlsx"
        output_dir = workspace / "recalculated"
        output_dir.mkdir()
        shutil.copy2(workbook_path, input_path)

        editable = load_workbook(input_path, data_only=False, keep_links=False)
        quotes = editable["Quotes"]
        for field, column in INPUT_COLUMNS.items():
            quotes[f"{column}4"] = inputs[field]
        editable.save(input_path)

        profile = (workspace / "lo-profile").resolve().as_uri()
        command = [
            soffice,
            "--headless",
            "--norestore",
            "--nodefault",
            "--nolockcheck",
            "--nofirststartwizard",
            f"-env:UserInstallation={profile}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_dir),
            str(input_path),
        ]
        timeout_seconds = int(os.getenv("PARITY_TIMEOUT_SECONDS", "45"))
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            check=False,
        )
        recalculated = output_dir / input_path.name
        if result.returncode != 0 or not recalculated.is_file():
            detail = (result.stderr or result.stdout or "LibreOffice did not produce a recalculated workbook.").strip()
            raise RuntimeError(detail[:500])

        calculated = load_workbook(recalculated, data_only=True, keep_links=False, read_only=True)
        try:
            row = next(calculated["Quotes"].iter_rows(min_row=4, max_row=4, values_only=False))
            return {
                field: _normalise_value(row[ord(column) - ord("A")].value)
                for field, column in OUTPUT_COLUMNS.items()
            }
        finally:
            calculated.close()


def _runtime_case(runtime: QuoteRuntime, scenario: ParityScenario) -> dict[str, Any]:
    quote = runtime.create_quote(QuoteCreateRequest.model_validate(scenario.inputs), persist=False).quotes[0]
    return {
        "unit_price": _normalise_value(quote.unit_price),
        "revenue": _normalise_value(quote.revenue),
        "cost": _normalise_value(quote.cost),
        "gross_margin": _normalise_value(quote.gross_margin),
        "approval_status": quote.approval_status,
    }


def _compare(workbook_result: dict[str, Any], runtime_result: dict[str, Any]) -> list[str]:
    diffs: list[str] = []
    for field in OUTPUT_COLUMNS:
        workbook_value = workbook_result.get(field)
        runtime_value = runtime_result.get(field)
        if field in TOLERANCES and workbook_value is not None and runtime_value is not None:
            if abs(float(workbook_value) - float(runtime_value)) > TOLERANCES[field]:
                diffs.append(f"{field}: workbook={workbook_value} runtime={runtime_value}")
        elif workbook_value != runtime_value:
            diffs.append(f"{field}: workbook={workbook_value} runtime={runtime_value}")
    return diffs


def run_parity(
    workbook_id: str,
    workbook_path: Path,
    blueprint: SystemBlueprint,
    scenarios: list[ParityScenario] | None = None,
) -> ParityRun:
    started_at = datetime.now(timezone.utc)
    started = time.perf_counter()
    selected = scenarios or golden_scenarios()
    runtime = QuoteRuntime(workbook_id, workbook_path, blueprint)
    soffice = find_soffice()
    blocked_reason = "LibreOffice recalculation is unavailable in this environment."
    results: list[ParityScenario] = []

    for scenario in selected:
        runtime_result: dict[str, Any] = {}
        try:
            runtime_result = _runtime_case(runtime, scenario)
        except (RuntimeValidationError, ValueError) as exc:
            results.append(
                scenario.model_copy(
                    update={
                        "runtime_result": {"error": str(exc)},
                        "status": ParityStatus.FAIL,
                        "diffs": [f"runtime: {exc}"],
                    }
                )
            )
            continue

        if not soffice:
            results.append(
                scenario.model_copy(
                    update={
                        "runtime_result": runtime_result,
                        "status": ParityStatus.BLOCKED,
                        "diffs": [blocked_reason],
                    }
                )
            )
            continue

        try:
            workbook_result = _recalculate_case(soffice, workbook_path, scenario.inputs)
            diffs = _compare(workbook_result, runtime_result)
            results.append(
                scenario.model_copy(
                    update={
                        "expected_outputs": workbook_result,
                        "workbook_result": workbook_result,
                        "runtime_result": runtime_result,
                        "status": ParityStatus.PASS if not diffs else ParityStatus.FAIL,
                        "diffs": diffs,
                    }
                )
            )
        except (OSError, RuntimeError, subprocess.TimeoutExpired) as exc:
            results.append(
                scenario.model_copy(
                    update={
                        "runtime_result": runtime_result,
                        "status": ParityStatus.BLOCKED,
                        "diffs": [f"workbook: {str(exc)[:500]}"],
                    }
                )
            )

    passed = sum(item.status is ParityStatus.PASS for item in results)
    failed = sum(item.status is ParityStatus.FAIL for item in results)
    blocked = sum(item.status is ParityStatus.BLOCKED for item in results)
    status = "fail" if failed else "blocked" if blocked else "pass"
    finished_at = datetime.now(timezone.utc)
    return ParityRun(
        run_id=f"parity-{uuid.uuid4().hex[:12]}",
        workbook_id=workbook_id,
        blueprint_version=blueprint.version,
        status=status,
        scenarios=results,
        total=len(results),
        passed=passed,
        failed=failed,
        blocked=blocked,
        started_at=started_at.isoformat(),
        finished_at=finished_at.isoformat(),
        duration_ms=int((time.perf_counter() - started) * 1000),
    )
