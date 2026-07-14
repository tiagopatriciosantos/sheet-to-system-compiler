"""Schema-driven quote runtime for the industrial demo workbook."""

from __future__ import annotations

import re
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

from app.domain.models import (
    BusinessRule,
    GeneratedAppResponse,
    QuoteCreateRequest,
    QuoteRecord,
    QuoteTransitionRequest,
    RuntimeClient,
    RuntimeDashboard,
    RuntimeProduct,
    SystemBlueprint,
    WorkflowSpec,
)
from app.runtime.storage import load_quotes, save_quotes


class RuntimeConfigurationError(ValueError):
    """Raised when the compiled app cannot prove that required inputs exist."""


class RuntimeValidationError(ValueError):
    """Raised when a quote input or transition violates the generated contract."""


def _key(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "_", ascii_value.lower()).strip("_")


def _as_float(value: Any, *, field: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise RuntimeConfigurationError(f"Workbook field {field} is not numeric.") from exc


def _rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise RuntimeConfigurationError(f"Required workbook sheet {sheet_name} was not found.")
    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(min_row=3, max_row=3, values_only=True), ())
    headers = [_key(value) for value in header_row]
    if not headers or not headers[0]:
        raise RuntimeConfigurationError(f"Sheet {sheet_name} has no header row at row 3.")
    records: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not any(value is not None for value in row):
            continue
        record = {header: value for header, value in zip(headers, row) if header}
        if record.get("id") or record.get("client_id") or record.get("product_id") or record.get("parameter"):
            records.append(record)
    return records


def _config(workbook: Any) -> dict[str, Any]:
    records = _rows(workbook, "Config")
    values: dict[str, Any] = {}
    for record in records:
        parameter = _key(record.get("parameter"))
        if parameter:
            values[parameter] = record.get("value")
    return values


def _approval_rule(blueprint: SystemBlueprint) -> BusinessRule | None:
    return next((rule for rule in blueprint.rules if rule.rule_type == "approval"), None)


def _workflow(blueprint: SystemBlueprint) -> WorkflowSpec | None:
    return next(iter(blueprint.workflows), None)


@dataclass(frozen=True)
class RuntimeContext:
    clients: dict[str, RuntimeClient]
    products: dict[str, RuntimeProduct]
    threshold: float
    default_max_discount: float
    rounding_decimals: int
    approval_operator: Literal["<", "<="]
    currency: str


class QuoteRuntime:
    """Executes only the small deterministic operation set represented in the blueprint."""

    def __init__(self, workbook_id: str, workbook_path: Path, blueprint: SystemBlueprint) -> None:
        self.workbook_id = workbook_id
        self.workbook_path = workbook_path
        self.blueprint = blueprint
        self._context = self._build_context()

    def _build_context(self) -> RuntimeContext:
        if not any(entity.name == "Quote" for entity in self.blueprint.entities):
            raise RuntimeConfigurationError("The blueprint has no Quote entity for the proposal runtime.")
        workbook = load_workbook(self.workbook_path, data_only=True, keep_links=False, read_only=True)
        client_records = _rows(workbook, "Clients")
        product_records = _rows(workbook, "Products")
        config = _config(workbook)
        clients = {
            str(record["client_id"]): RuntimeClient(
                id=str(record["client_id"]),
                name=str(record.get("client_name") or record["client_id"]),
                max_discount=_as_float(record.get("max_discount"), field="Clients.max_discount"),
            )
            for record in client_records
        }
        products = {
            str(record["product_id"]): RuntimeProduct(
                id=str(record["product_id"]),
                sku=str(record.get("sku") or ""),
                description=str(record.get("description") or record["product_id"]),
                unit_cost=_as_float(record.get("unit_cost"), field="Products.unit_cost"),
                base_price=_as_float(record.get("base_price"), field="Products.base_price"),
            )
            for record in product_records
        }
        approval_rule = _approval_rule(self.blueprint)
        if approval_rule is None:
            raise RuntimeConfigurationError("The blueprint has no approval rule for the quote runtime.")
        threshold_value = config.get("margin_approval_threshold") or config.get("marginapprovalthreshold")
        if threshold_value is None:
            raise RuntimeConfigurationError("Config.MarginApprovalThreshold is required by the approval runtime.")
        operator: Literal["<", "<="] = "<=" if "<=" in (approval_rule.expression or "") else "<"
        return RuntimeContext(
            clients=clients,
            products=products,
            threshold=_as_float(threshold_value, field="Config.MarginApprovalThreshold"),
            default_max_discount=_as_float(
                config.get("default_max_discount", config.get("defaultmaxdiscount", 1)),
                field="Config.DefaultMaxDiscount",
            ),
            rounding_decimals=int(config.get("rounding_decimals", config.get("roundingdecimals", 2))),
            approval_operator=operator,
            currency=str(config.get("currency") or "EUR"),
        )

    def snapshot(self) -> GeneratedAppResponse:
        quotes = load_quotes(self.workbook_id)
        return self._response(quotes)

    def create_quote(self, request: QuoteCreateRequest) -> GeneratedAppResponse:
        client = self._context.clients.get(request.client_id)
        if client is None:
            raise RuntimeValidationError(f"Unknown client: {request.client_id}.")
        product = self._context.products.get(request.product_id)
        if product is None:
            raise RuntimeValidationError(f"Unknown product: {request.product_id}.")
        allowed_discount = client.max_discount
        if request.discount > allowed_discount:
            raise RuntimeValidationError(
                f"Discount exceeds the client policy ({allowed_discount:.2%})."
            )

        revenue = request.quantity * product.base_price * (1 - request.discount)
        cost = request.quantity * product.unit_cost
        margin = 0.0 if revenue == 0 else (revenue - cost) / revenue
        needs_approval = (
            margin <= self._context.threshold
            if self._context.approval_operator == "<="
            else margin < self._context.threshold
        )
        status = "NEEDS_APPROVAL" if needs_approval else "AUTO_APPROVED"
        reason = (
            f"Gross margin {margin:.2%} is below the configured threshold {self._context.threshold:.2%}."
            if needs_approval
            else f"Gross margin {margin:.2%} meets the configured threshold {self._context.threshold:.2%}."
        )
        decimals = self._context.rounding_decimals
        quote = QuoteRecord(
            id=f"Q-{uuid.uuid4().hex[:10].upper()}",
            client_id=client.id,
            product_id=product.id,
            quantity=request.quantity,
            discount=request.discount,
            unit_price=round(product.base_price, decimals),
            revenue=round(revenue, decimals),
            cost=round(cost, decimals),
            gross_margin=margin,
            approval_status=status,
            evidence_reason=reason,
            created_at=datetime.now(timezone.utc).isoformat(),
        )
        quotes = [*load_quotes(self.workbook_id), quote]
        save_quotes(self.workbook_id, quotes)
        return self._response(quotes)

    def transition_quote(self, quote_id: str, request: QuoteTransitionRequest) -> GeneratedAppResponse:
        quotes = load_quotes(self.workbook_id)
        for index, quote in enumerate(quotes):
            if quote.id != quote_id:
                continue
            if quote.approval_status != "NEEDS_APPROVAL":
                raise RuntimeValidationError("Only quotes needing approval can be transitioned.")
            workflow = _workflow(self.blueprint)
            if workflow and not any(
                transition.get("from") == quote.approval_status
                and transition.get("to") == request.target_status
                for transition in workflow.transitions
            ):
                raise RuntimeValidationError(
                    f"Transition {quote.approval_status} -> {request.target_status} is not in the blueprint workflow."
                )
            quotes[index] = quote.model_copy(
                update={
                    "approval_status": request.target_status,
                    "transition_note": request.note or None,
                }
            )
            save_quotes(self.workbook_id, quotes)
            return self._response(quotes)
        raise RuntimeValidationError(f"Quote {quote_id} was not found.")

    def _response(self, quotes: list[QuoteRecord]) -> GeneratedAppResponse:
        approved = sum(quote.approval_status in {"AUTO_APPROVED", "APPROVED"} for quote in quotes)
        return GeneratedAppResponse(
            workbook_id=self.workbook_id,
            blueprint_version=self.blueprint.version,
            unresolved_items=self.blueprint.unresolved_items,
            workflow=_workflow(self.blueprint),
            clients=sorted(self._context.clients.values(), key=lambda item: item.id),
            products=sorted(self._context.products.values(), key=lambda item: item.id),
            quotes=list(reversed(quotes)),
            dashboard=RuntimeDashboard(
                total_quotes=len(quotes),
                needs_approval=sum(quote.approval_status == "NEEDS_APPROVAL" for quote in quotes),
                approved_quotes=approved,
                rejected_quotes=sum(quote.approval_status == "REJECTED" for quote in quotes),
                total_revenue=round(sum(quote.revenue for quote in quotes), self._context.rounding_decimals),
            ),
        )
