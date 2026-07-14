"""Deterministic workbook -> WorkbookIR extraction for the Fase 1 X-Ray."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from app.domain.models import (
    ConditionalFormatRule,
    EvidenceRef,
    FormulaCell,
    SheetVisibility,
    WorkbookIR,
    WorkbookSheet,
    WorkbookValidation,
)


CELL_REFERENCE_RE = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ ]*))!)?(\$?[A-Z]{1,3}\$?\d+)"
)
UNSUPPORTED_FORMULA_RE = re.compile(r"\b(?:WEBSERVICE|RTD|CUBE[A-Z]+|SQL\.REQUEST)\b", re.IGNORECASE)


def _visibility(sheet_state: str) -> SheetVisibility:
    return {
        "visible": SheetVisibility.VISIBLE,
        "hidden": SheetVisibility.HIDDEN,
        "veryHidden": SheetVisibility.VERY_HIDDEN,
    }.get(sheet_state, SheetVisibility.VISIBLE)


def _safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _table_names(sheet: Any) -> list[str]:
    return [name for name, _table in sheet.tables.items()]


def _named_ranges(workbook: Any) -> list[str]:
    return [name for name, _defined_name in workbook.defined_names.items()]


def _formula_dependencies(sheet_name: str, formula: str) -> list[dict[str, str]]:
    dependencies: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for match in CELL_REFERENCE_RE.finditer(formula):
        referenced_sheet = match.group(1) or match.group(2) or sheet_name
        address = match.group(3).replace("$", "")
        key = (referenced_sheet, address)
        if key not in seen:
            seen.add(key)
            dependencies.append({"sheet": referenced_sheet, "address": address})
    return dependencies


def _conditional_formats(sheet: Any, evidence: list[EvidenceRef]) -> list[ConditionalFormatRule]:
    extracted: list[ConditionalFormatRule] = []
    for sqref in sheet.conditional_formatting:
        rules = sheet.conditional_formatting[sqref]
        for index, rule in enumerate(rules, start=1):
            evidence_id = f"{sheet.title}!conditional-format-{len(extracted) + 1}"
            evidence.append(
                EvidenceRef(
                    id=evidence_id,
                    source_type="range",
                    sheet=sheet.title,
                    address=str(sqref),
                    excerpt=f"Conditional formatting rule {index}",
                )
            )
            formulas = getattr(rule, "formula", None) or []
            extracted.append(
                ConditionalFormatRule(
                    range=str(sqref),
                    rule_type=type(rule).__name__,
                    operator=getattr(rule, "operator", None),
                    formula=" | ".join(str(item) for item in formulas) or None,
                    evidence_refs=[evidence_id],
                )
            )
    return extracted


def _validations(sheet: Any, evidence: list[EvidenceRef]) -> list[WorkbookValidation]:
    extracted: list[WorkbookValidation] = []
    for index, validation in enumerate(sheet.data_validations.dataValidation, start=1):
        evidence_id = f"{sheet.title}!validation-{index}"
        target = str(validation.sqref)
        evidence.append(
            EvidenceRef(
                id=evidence_id,
                source_type="range",
                sheet=sheet.title,
                address=target,
                excerpt=f"Data validation: {validation.type or 'unknown'}",
            )
        )
        extracted.append(
            WorkbookValidation(
                range=target,
                validation_type=validation.type,
                operator=validation.operator,
                formula1=str(validation.formula1) if validation.formula1 is not None else None,
                formula2=str(validation.formula2) if validation.formula2 is not None else None,
                evidence_refs=[evidence_id],
            )
        )
    return extracted


def extract_workbook(path: Path, workbook_id: str, filename: str, sha256: str) -> WorkbookIR:
    """Extract formulas, cached values, structure and evidence without executing workbook code."""

    formulas_workbook = load_workbook(path, data_only=False, read_only=False, keep_links=False)
    values_workbook = load_workbook(path, data_only=True, read_only=False, keep_links=False)
    evidence: list[EvidenceRef] = []
    dependencies: list[dict[str, str]] = []
    unsupported: list[str] = []
    declared_unsupported_markers: set[str] = set()
    named_ranges = _named_ranges(formulas_workbook)

    sheets: list[WorkbookSheet] = []
    for formula_sheet in formulas_workbook.worksheets:
        values_sheet = values_workbook[formula_sheet.title]
        sheet_evidence_ids: set[str] = set()
        formula_cells: list[FormulaCell] = []
        constants: list[str] = []
        sheet_dependencies: list[dict[str, str]] = []
        warnings: list[str] = []

        for row in formula_sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell) or cell.value is None:
                    continue
                evidence_id = f"{formula_sheet.title}!{cell.coordinate}"
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cached_value = values_sheet[cell.coordinate].value
                    evidence.append(
                        EvidenceRef(
                            id=evidence_id,
                            source_type="formula",
                            sheet=formula_sheet.title,
                            address=cell.coordinate,
                            excerpt=cell.value,
                        )
                    )
                    formula_cells.append(
                        FormulaCell(
                            address=cell.coordinate,
                            formula=cell.value,
                            cached_value=_safe_value(cached_value),
                            evidence_refs=[evidence_id],
                        )
                    )
                    for dependency in _formula_dependencies(formula_sheet.title, cell.value):
                        dependency_record = {
                            "source_sheet": formula_sheet.title,
                            "source_address": cell.coordinate,
                            "target_sheet": dependency["sheet"],
                            "target_address": dependency["address"],
                        }
                        if dependency_record not in dependencies:
                            dependencies.append(dependency_record)
                        if dependency_record not in sheet_dependencies:
                            sheet_dependencies.append(dependency_record)
                    if UNSUPPORTED_FORMULA_RE.search(cell.value):
                        unsupported.append(f"Unsupported formula in {formula_sheet.title}!{cell.coordinate}")
                else:
                    constants.append(cell.coordinate)
                    if isinstance(cell.value, str) and (
                        "unsupportedfeature" in cell.value.lower()
                        or "powerquery" in cell.value.lower()
                        or "power query" in cell.value.lower()
                    ):
                        declared_unsupported_markers.add("PowerQueryRefresh")
                    if len(constants) <= 20:
                        evidence.append(
                            EvidenceRef(
                                id=evidence_id,
                                source_type="cell",
                                sheet=formula_sheet.title,
                                address=cell.coordinate,
                                excerpt=str(cell.value)[:160],
                            )
                        )
                        sheet_evidence_ids.add(evidence_id)

        visibility = _visibility(formula_sheet.sheet_state)
        if visibility != SheetVisibility.VISIBLE:
            warnings.append("Hidden sheet may contain business parameters and requires explicit review.")

        sheet_validations = _validations(formula_sheet, evidence)
        sheet_conditional_formats = _conditional_formats(formula_sheet, evidence)
        if sheet_conditional_formats:
            warnings.append("Conditional formatting was extracted as evidence; its business meaning is not assumed.")

        sheets.append(
            WorkbookSheet(
                name=formula_sheet.title,
                visibility=visibility,
                max_row=formula_sheet.max_row,
                max_column=formula_sheet.max_column,
                tables=_table_names(formula_sheet),
                named_ranges=named_ranges,
                merged_ranges=[str(item) for item in formula_sheet.merged_cells.ranges],
                formula_cells=formula_cells,
                constants_of_interest=constants[:20],
                validations=sheet_validations,
                conditional_formats=sheet_conditional_formats,
                warnings=warnings,
            )
        )

    external_links = [str(link) for link in getattr(formulas_workbook, "_external_links", [])]
    unsupported.extend(f"Declared unsupported feature: {marker}" for marker in sorted(declared_unsupported_markers))
    if external_links:
        unsupported.append("External workbook links detected; refresh is blocked in the MVP.")

    unique_unsupported = list(dict.fromkeys(unsupported))
    return WorkbookIR(
        workbook_id=workbook_id,
        sha256=sha256,
        filename=filename,
        sheets=sheets,
        formula_dependencies=dependencies,
        external_links=external_links,
        unsupported_features=unique_unsupported,
        evidence=evidence,
    )
